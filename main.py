import openpyxl 

inv_file = openpyxl.load_workbook("inventory.xlsx")
products_list = inv_file["Sheet1"]

# Calculate how many product we have per suplier and list of the supliers with respective names

products_per_suplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

for product_row in range(2, products_list.max_row + 1):
    supplier_name = products_list.cell(product_row, 4).value
    inventory = products_list.cell(product_row, 2).value
    price = products_list.cell(product_row, 3).value
    product_num = products_list.cell(product_row, 1).value
    inventory_price = products_list.cell(product_row, 5)


    # Calculation for number of products pwe suppliers
    if supplier_name in products_per_suplier:
        curent_number_of_products = products_per_suplier.get(supplier_name)
        products_per_suplier[supplier_name] = curent_number_of_products + 1
    else:
        products_per_suplier[supplier_name] = 1


# Calculate total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        curent_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = curent_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

# Print all the products that have inventory less then 10
    if inventory <  10:
        products_under_10_inv[int(product_num)] = int(inventory)

# Add values for total inventory price
    inventory_price.value = inventory * price
inv_file.save("inventory_with_total_value.xlsx")